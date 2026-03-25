import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils import get_column_letter

# 強制清除任何可能殘留的 CSS 隱藏樣式
st.markdown("<style>button { visibility: visible !important; }</style>", unsafe_allow_html=True)

st.set_page_config(page_title="醫療帳務資料合併工具", layout="wide")
st.title("🏥 醫療帳務資料合併工具")

# --- 初始化 Session State ---
if 'processed_output' not in st.session_state:
    st.session_state.processed_output = None
if 'detailed_records' not in st.session_state:
    st.session_state.detailed_records = []
if 'target_date_str' not in st.session_state:
    st.session_state.target_date_str = None
if 'data_pool' not in st.session_state:
    st.session_state.data_pool = {}
for key in ['audit_sheet1', 'audit_sheet2', 'audit_sheet3', 'audit_sheet45', 'hp_details']:
    if getattr(st.session_state, key, None) is None:
        st.session_state[key] = []

# --- 檔案上傳區 ---
uploaded_files = st.file_uploader("請同時選擇或拖入「115年度明細表」與「每日來源資料」兩個檔案", type=["xlsx", "xlsm"], accept_multiple_files=True)

template_file = None
day_file = None

if uploaded_files:
    for f in uploaded_files:
        try:
            # 讀取 Sheet 清單來判斷身分
            xls = pd.ExcelFile(f)
            sheet_names = xls.sheet_names
            # 只要有「代號表」或「工作表1」就是來源檔
            if "代號表" in sheet_names or "工作表1" in sheet_names:
                day_file = f
            # 只要有「115」開頭的分頁就是模板檔
            elif any(s.startswith("115") for s in sheet_names):
                template_file = f
        except Exception:
            continue

# 只要偵測到兩個檔案都到齊，就顯示開始按鈕
if template_file and day_file:
    st.success(f"📁 檔案已就緒：\n- 模板：{template_file.name}\n- 來源：{day_file.name}")
    
    if st.button("🚀 開始精準合併並對帳", type="primary"):
        with st.spinner("正在執行對帳座標運算..."):
            try:
                # 1. 讀取代號表
                df_codes = pd.read_excel(day_file, sheet_name="代號表")
                code_dict = {}
                for _, row in df_codes.iterrows():
                    name = str(row.iloc[0]).strip()
                    for i in range(1, len(row)):
                        if pd.notna(row.iloc[i]):
                            val = str(row.iloc[i]).split('.')[0]
                            c = val.zfill(2) if val.isdigit() and len(val) < 3 else val
                            code_dict[c] = name

                # 2. 建立資料彙整池
                st.session_state.data_pool = {}
                st.session_state.detailed_records = []
                st.session_state.target_date_str = None
                st.session_state.audit_sheet1 = []
                st.session_state.audit_sheet2 = []
                st.session_state.audit_sheet3 = []
                st.session_state.audit_sheet45 = []
                st.session_state.hp_details = []
                
                def collect_data(date_obj, col, val, reason, name):
                    if val == 0: return
                    d_str = date_obj.strftime('%Y-%m-%d')
                    key = (d_str, col)
                    old_v, _, _ = st.session_state.data_pool.get(key, (0.0, "", ""))
                    st.session_state.data_pool[key] = (old_v + val, reason, name)
                    st.session_state.detailed_records.append({
                        "日期": d_str, "醫師/對象": name, "欄位編號": col, "項目內容": reason, "金額": val
                    })

                # 座標地圖 (VBA 原版座標)
                opd_stu = {'李':(40,41,42),'珩':(43,44,45),'芳':(46,47,48),'東':(49,50,51),'澍':(52,53,54),'張明揚':(55,56,57),'李建南':(58,59,60),'影像':(64,65,66)}
                opd_no_stu = {'鄭':61, '許越涵':62, '陳思宇':63}
                birth_map = {'李':76,'珩':77,'芳':78,'東':79,'澍':80,'李建南':81,'張明揚':82,'鄭':83,'陳思宇':84}
                room_map = {'李':85,'珩':86,'芳':87,'東':88,'澍':89,'李建南':90,'張明揚':91,'鄭':92,'陳思宇':93,'林慧雯':94}
                nurs_map = {'李':115,'珩':116,'芳':117,'東':118,'澍':119,'李建南':120,'張明揚':121,'林慧雯':122}

                def safe_num(v):
                    try: return float(v) if pd.notna(v) else 0.0
                    except: return 0.0

                day_xls = pd.ExcelFile(day_file)
                all_sheets = day_xls.sheet_names

                # 3. 工作表1 (門診)
                if "工作表1" in all_sheets:
                    df1 = pd.read_excel(day_file, sheet_name="工作表1", header=None, skiprows=1)
                    for _, row in df1.iterrows():
                        dt = pd.to_datetime(row.iloc[0], errors='coerce')
                        if pd.isna(dt): continue
                        if st.session_state.target_date_str is None or dt.strftime('%Y-%m-%d') > st.session_state.target_date_str:
                            st.session_state.target_date_str = dt.strftime('%Y-%m-%d')
                        full_code = str(row.iloc[1]).strip()
                        c = full_code.split('.')[0].zfill(2)
                        name = code_dict.get(c)
                        val = safe_num(row.iloc[16]) - safe_num(row.iloc[4]) - safe_num(row.iloc[5])
                        
                        reg_fee = safe_num(row.iloc[4])
                        copay_fee = safe_num(row.iloc[5])
                        
                        st.session_state.audit_sheet1.append({
                            "日期": dt.strftime('%Y-%m-%d'),
                            "對象": name if name else "未知",
                            "小計": safe_num(row.iloc[16]),
                            "掛號": reg_fee,
                            "部分負擔": copay_fee,
                            "門診金額": val
                        })
                        
                        # 將掛號費與部分負擔累加，分別寫入 AE(31) 與 AF(32) 欄
                        collect_data(dt, 31, reg_fee, "掛號費", "門診總計")
                        collect_data(dt, 32, copay_fee, "部分負擔", "門診總計")
                        
                        if name == '兒sona': collect_data(dt, 71, val, "兒sona", "兒sona")
                        elif name == '兒科': collect_data(dt, 70, val, "兒科", "兒科")
                        elif name == '外賣': collect_data(dt, 124, val, "外賣", "外賣")
                        elif name == '哺乳諮詢': collect_data(dt, 67, val, "哺乳諮詢", "哺乳諮詢")
                        elif name == '營養諮詢': collect_data(dt, 68, val, "營養諮詢", "營養諮詢")
                        elif name == '助產諮詢': collect_data(dt, 69, val, "助產諮詢", "助產諮詢")
                        elif name in opd_no_stu: collect_data(dt, opd_no_stu[name], val, "門診", name)
                        elif name in opd_stu:
                            s = str(row.iloc[2]).strip().upper()
                            s_idx = 0 if s=='S' else (1 if s=='T' else 2)
                            label = {'S':'早', 'T':'午', 'U':'晚'}.get(s, s)
                            collect_data(dt, opd_stu[name][s_idx], val, f"OPD({label})", name)

                # 4. 工作表2 (出院與生產實收)
                if "工作表2" in all_sheets:
                    df2 = pd.read_excel(day_file, sheet_name="工作表2", header=None, skiprows=1)
                    hp_agg = {}
                    for _, row in df2.iterrows():
                        dt = pd.to_datetime(row.iloc[0], errors='coerce')
                        if pd.isna(dt): continue
                        c = str(row.iloc[2]).strip().split('.')[0].zfill(2)
                        name = code_dict.get(c, "其他")
                        iAnes, iRoom, iBirth, iMat, iPre, iFood = safe_num(row.iloc[7]), safe_num(row.iloc[8]), safe_num(row.iloc[9]), safe_num(row.iloc[10]), safe_num(row.iloc[11]), safe_num(row.iloc[12])
                        
                        if name in room_map:
                            collect_data(dt, room_map[name], iRoom, "病房費", name)
                            collect_data(dt, room_map[name]+10, iMat, "材料費", name)
                            collect_data(dt, room_map[name]+20, iFood, "伙食費", name)
                            st.session_state.audit_sheet2.append({
                                "日期": dt.strftime('%Y-%m-%d'), "對象": name, "項目": "一般出院",
                                "明細": f"病房:{iRoom}, 材料:{iMat}, 伙食:{iFood}", "金額": iRoom + iMat + iFood
                            })
                        
                        if iPre >= 0:
                            birth_total = iAnes + iBirth + iPre
                            if birth_total != 0 and name in birth_map:
                                collect_data(dt, birth_map[name], birth_total, "生產實收(麻+產+預)", name)
                                st.session_state.audit_sheet2.append({
                                    "日期": dt.strftime('%Y-%m-%d'), "對象": name, "項目": "生產實收",
                                    "明細": f"麻醉:{iAnes}, 產費:{iBirth}, 預收:{iPre}", "金額": birth_total
                                })
                        else:
                            hp_val = abs(iPre) - iAnes - iBirth
                            d_str = dt.strftime('%Y-%m-%d')
                            hp_agg[d_str] = hp_agg.get(d_str, 0.0) + hp_val
                            patient_name = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else "未知"
                            st.session_state.audit_sheet2.append({
                                "日期": d_str, "對象": name, "項目": "HP結算(單筆)",
                                "明細": f"Abs(預收:{iPre}) - 麻醉:{iAnes} - 產費:{iBirth}", "金額": hp_val
                            })
                            st.session_state.hp_details.append({
                                "日期": d_str,
                                "產婦姓名": patient_name,
                                "HP結算金額": hp_val
                            })
                    
                    for d_str, total in hp_agg.items():
                        if total != 0: 
                            collect_data(datetime.strptime(d_str, '%Y-%m-%d'), 224, total, "HP結算", "總計")
                            st.session_state.audit_sheet2.append({
                                "日期": d_str, "對象": "全部對象", "項目": "HP結算(單日加總)",
                                "明細": "當日所有預收款負數相加之總額", "金額": total
                            })

                # 5. 工作表3 (嬰兒室)
                if "工作表3" in all_sheets:
                    df3 = pd.read_excel(day_file, sheet_name="工作表3", header=None, skiprows=1)
                    for _, row in df3.iterrows():
                        dt = pd.to_datetime(row.iloc[0], errors='coerce')
                        if pd.isna(dt): continue
                        c = str(row.iloc[2]).strip().split('.')[0].zfill(2)
                        val = safe_num(row.iloc[6])
                        name = code_dict.get(c)
                        if name in nurs_map: 
                            collect_data(dt, nurs_map[name], val, "嬰兒室", name)
                            st.session_state.audit_sheet3.append({
                                "日期": dt.strftime('%Y-%m-%d'), "對象": name, "小計金額": val
                            })

                # 6 & 7. 欠款與還款
                for sheet, col_keyword, label, target_col in [("工作表4", "未收額", "今日欠款", 135), ("工作表5", "還款金額", "今日還款", 123)]:
                    if sheet in all_sheets:
                        tmp = pd.read_excel(day_file, sheet_name=sheet)
                        dt_col = next((c for c in tmp.columns if '日期' in str(c)), tmp.columns[0])
                        val_col = next((c for c in tmp.columns if col_keyword in str(c)), None)
                        if val_col:
                            for _, row in tmp.iterrows():
                                dt = pd.to_datetime(row[dt_col], errors='coerce')
                                if pd.isna(dt): continue
                                collect_data(dt, target_col, safe_num(row[val_col]), label, "總計")
                                st.session_state.audit_sheet45.append({
                                    "類別": label, "日期": dt.strftime('%Y-%m-%d'), "來源金額": safe_num(row[val_col])
                                })

                # --- 寫入 Excel ---
                template_file.seek(0)
                wb = load_workbook(template_file)
                for (d_str, col), (val, reason, name) in st.session_state.data_pool.items():
                    dt = datetime.strptime(d_str, '%Y-%m-%d')
                    m_key = f"115{dt.month:02d}"
                    if m_key in wb.sheetnames: wb[m_key].cell(row=dt.day + 3, column=col).value = val

                out = io.BytesIO()
                wb.save(out)
                st.session_state.processed_output = out.getvalue()
                st.success("✅ 處理完成！")

            except Exception as e:
                st.error(f"發生錯誤: {e}")
                st.exception(e)

# --- 顯示結果區域 ---
if st.session_state.processed_output is not None:
    st.divider()
    st.download_button(label="💾 下載結果檔案", data=st.session_state.processed_output, file_name=f"{datetime.now().strftime('%Y%m%d')}_財務對帳版.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

    if st.session_state.target_date_str:
        st.header(f"🔍 獨立對帳面板 (工作表原始數據 - {st.session_state.target_date_str})")
        tab1, tab2, tab3, tab4 = st.tabs(["工作表1 (門診)", "工作表2 (出院與生產)", "工作表3 (嬰兒室)", "工作表4&5 (欠還款)"])
        
        with tab1:
            df1 = [r for r in st.session_state.audit_sheet1 if r.get('日期') == st.session_state.target_date_str]
            if df1:
                st.dataframe(pd.DataFrame(df1), use_container_width=True, hide_index=True)
            else:
                st.info(f"當日({st.session_state.target_date_str})無 工作表1 資料")

        with tab2:
            df2 = [r for r in st.session_state.audit_sheet2 if r.get('日期') == st.session_state.target_date_str]
            if df2:
                st.dataframe(pd.DataFrame(df2), use_container_width=True, hide_index=True)
            else:
                st.info(f"當日({st.session_state.target_date_str})無 工作表2 資料")

        with tab3:
            df3 = [r for r in st.session_state.audit_sheet3 if r.get('日期') == st.session_state.target_date_str]
            if df3:
                st.dataframe(pd.DataFrame(df3), use_container_width=True, hide_index=True)
            else:
                st.info(f"當日({st.session_state.target_date_str})無 工作表3 資料")

        with tab4:
            df4 = [r for r in st.session_state.audit_sheet45 if r.get('日期') == st.session_state.target_date_str]
            if df4:
                st.dataframe(pd.DataFrame(df4), use_container_width=True, hide_index=True)
            else:
                st.info(f"當日({st.session_state.target_date_str})無 欠還款 資料")
                
        st.divider()
        st.header(f"📊 詳細對帳單 ({st.session_state.target_date_str})")
        day_pool = {k: v for k, v in st.session_state.data_pool.items() if k[0] == st.session_state.target_date_str}
        if day_pool:
            final_list = []
            for (d, c), (v, r, n) in day_pool.items():
                col_letter = get_column_letter(c)
                final_list.append({"醫師/對象": n, "項目名稱": r, "Excel欄位": f"{col_letter} ({c})", "金額": v, "編號": c})
            display_df = pd.DataFrame(final_list).sort_values(by=['編號', '醫師/對象'])
            display_df['金額'] = display_df['金額'].apply(lambda x: f"{x:,.0f}")
            st.dataframe(display_df[['醫師/對象', '項目名稱', 'Excel欄位', '金額']], use_container_width=True, hide_index=True)
            st.info("💡 提示：表格已鎖定，您可以放心地點擊下載按鈕。")
        else:
            st.warning("當日無異動。")
            
        st.divider()
        st.header(f"🛏️ 住院預收款(HP結算)明細 ({st.session_state.target_date_str})")
        hp_records = [r for r in st.session_state.hp_details if r.get('日期') == st.session_state.target_date_str]
        if hp_records:
            hp_df = pd.DataFrame(hp_records)
            hp_df['HP結算金額'] = hp_df['HP結算金額'].apply(lambda x: f"{x:,.0f}")
            st.dataframe(hp_df[['產婦姓名', 'HP結算金額']], use_container_width=True, hide_index=True)
        else:
            st.info("當日無住院預收款結算資料。")
elif uploaded_files and (template_file is None or day_file is None):
    st.warning("請確保同時上傳了「115年度明細表」與「每日來源資料 (day.xlsx)」兩個檔案。")